import os
import datetime
import requests
import boto3
import logging
import openpyxl as XL
import xml.etree.ElementTree as ET

BUCKET = boto3.resource('s3').Bucket('ebayreports')

LOG = logging.getLogger()
LOG.setLevel(logging.ERROR)

STORE_NAMES = os.environ['storeNames'].split(',')
KEY = os.environ['key']

URL = 'https://api.ebay.com/ws/api.dll'
	
HEADERS = {
	'Content-Type' : 'text/xml',
	'X-EBAY-API-COMPATIBILITY-LEVEL' : '1081',
	'X-EBAY-API-CALL-NAME' : 'GetSellerEvents',
	'X-EBAY-API-SITEID' : '0'
	}
	
REPORT_FIELDS = [
	'itemid',
	'status',
	'title',
	'price',
	'last_price',
	'price_difference',
	'url'
	]

PRE = '{urn:ebay:apis:eBLBaseComponents}'
	
TODAY = datetime.datetime.now() - datetime.timedelta(hours=8)
TODAY_STRING = TODAY.strftime('%m-%d-%Y %I:%M%p')

DT_MOD = 'Mod'
DT_NEW = 'Start'
DT_REM = 'End'

FN_LASTRUN = 'lastrun.txt'
FN_DATA = 'data.xlsx'
FN_REPORT = 'report.xlsx'

ST_INC = "INCREASED"
ST_RED = "REDUCED"
ST_NEW = "NEW"
ST_END = "REMOVED"

def P(str):
	return f'{PRE}{str}'
	
def getXML(storeName, fromDate, toDate, dateType):
	return f"""
<?xml version="1.0" encoding="utf-8"?>
<GetSellerEventsRequest xmlns="urn:ebay:apis:eBLBaseComponents">
	<RequesterCredentials>
    <eBayAuthToken>{apiKey}</eBayAuthToken>
  </RequesterCredentials>
  <{dateType}TimeFrom>{fromDate}</{dateType}TimeFrom>
  <{dateType}TimeTo>{toDate}</{dateType}TimeTo>
  <UserID>{storeName}</UserID>
  <DetailLevel>ReturnAll</DetailLevel>
  <OutputSelector>ItemID</OutputSelector>
  <OutputSelector>ListingDetails</OutputSelector>
  <OutputSelector>SellingStatus</OutputSelector>
</GetSellerEventsRequest>"""

def getLastRunTime(storeName):
	try:
		timestring = bucket.Object(f"{storeName}/{FN_LASTRUN}").get()['Body'].read().decode('utf-8')
		return datetime.datetime.strptime(timestring, '%Y-%m-%dT%H:%M:%S.%fZ')
	except Exception as err:
		LOG.error(f"[{storeName}] Error reading {FN_LASTRUN}: {err}")
		return None
		
def getLastRunData(storeName):
	try:
		bucket.download_file(f"{storeName}/{FN_DATA}", "/tmp/{storeName}/_{FN_DATA}")
		lastData_wb = XL.load_workbook(filename = "/tmp/{storeName}/_{FN_DATA}", read_only=True)
		lastData_ws = lastData_wb['Sheet']
		
		items = {}
		
		for row in lastData_ws.rows:
			items[row[0].value] = row[1].value
			
		return items
	except Exception as err:
		LOG.error(f"[{storeName}] Error reading {FN_DATA}: {err}")
		return None
		
def getListings(storeName, previousTimestamp, dateType):
	body = getXML(storeName, previousTimestamp, TODAY, dateType)
	
	try:
		res = requests.post(URL, data=body, headers=HEADERS)
		root = ET.fromstring(res.content)
		itemList = root.find(P('ItemArray'))
		return itemList
	except Exception as err:
		LOG.error(f"[{storeName}] Error getting {dateType} listings: {err}")
		return None
	
def putToS3(remoteName, localName):
	BUCKET.Object(remoteName).put(Body=open(localName, 'rb'))
		
def main(event, context):
	
	for storeName in STORE_NAMES:
		data = getLastRunData(storeName)
		# Move to next store if we can't retrieve data
		if not previousData:
			continue
			
		previousTimestamp = getLastRunTime(storeName)
		# Move to next store if we can't retrieve last run timestamp
		if not previousTimestamp:
			continue

		report = []
		
		modListings = getListings(storeName, previousTimestamp, DT_MOD)
		newListings = getListings(storeName, previousTimestamp, DT_NEW)
		endListings = getListings(storeName, previousTimestamp, DT_REM)
		
		### Process modified listings ###
		if modListings:
			# loop through modListings
			for eachItem in modListings:
				# get bare necessity data
				itemID = eachItem.find(P('ItemID')).text
				curPrice = eachItem.find(P('SellingStatus')).find(P('CurrentPrice')).text
				lastPrice = data[itemID]
			
				# get price difference
				priceDiff = float(curPrice) - float(lastPrice)
				
				# if different...
				if priceDiff != 0:
					title = eachItem.find(P('Title')).text
					url = eachItem.find(P('ListingDetails').find('ViewItemURL').text
					
					if priceDiff > 0:
						status = ST_INC
					else:
						status = ST_RED
					
					# Add to report
					report.append({
						'itemid' : itemID,
						'status' : status,
						'title' : title,		
						'price' : curPrice,
						'last_price' : lastPrice,
						'price_difference' : priceDiff,
						'url' : url})
						
					# Update data
					data[itemID] = curPrice
		### End processing modified listings
		
		### Process new listings ###
		if newListings:
			# loop through newListings
			for eachItem in newListings:
				# Get fields
				itemID = eachItem.find(P('ItemID')).text
				curPrice = eachItem.find(P('SellingStatus')).find(P('CurrentPrice')).text
				title = eachItem.find(P('Title')).text
				url = eachItem.find(P('ListingDetails').find(P('ViewItemURL')).text
				
				# Add to report
				report.append({
					'itemid' : itemID,		
					'status' : ST_NEW,
					'title' : title,		
					'price' : curPrice,	
					'last_price' : '',
					'price_difference' : '',
					'url' : url})		
					
				# Add to data
				data[itemID] = curPrice
		### End processing new listings ###
		
		### Process ended listings ###
		if endListings:
			# Loop through endListings
			for eachItem in endListings:
				# Get fields
				itemID = eachItem.find(P('ItemID')).text
				curPrice = eachItem.find(P('SellingStatus')).find(P('CurrentPrice')).text
				title = eachItem.find(P('Title')).text
				
				# Add to report
				report.append({
					'itemid' : itemID,
					'status' : ST_END,
					'title' : title,
					'price' : curPrice,
					'last_price' : data[itemID],
					'url' : ''})
					
				# Remove from data
				del data[itemID]
		### End processing ended listings ###
				
		### Write timestamp ###
		LOG.info(f"[{storeName}] Writing LASTRUN...")
		with open(f"/tmp/{storeName}/{FN_LASTRUN}", 'w', newline='') as timeFile:
			timeFile.write(f"{TODAY}")
			
		putToS3(f"{storeName}/{FN_LASTRUN}", f"/tmp/{storeName}/{FN_LASTRUN}")
		### End writing timestamp  ###
	
		### Write data ###
		LOG.info(f"[{storeName}] Writing DATA...")
		wb_data = XL.Workbook()
		ws_data = wb_data.active
		
		for itemid in data:
			ws_data.append([itemid, data[itemid])
			
		wb_data.save("/tmp/{storeName}/{FN_DATA}")
		putToS3(f"{storeName}/{FN_DATA}", f"/tmp/{storeName}/{FN_DATA}")
		### End writing data ###
				
		### Write report ###
		if report:
			LOG.info(f"[{storeName}] Writing REPORT...")
			wb = XL.Workbook()
			ws = wb.active
			ws.append(REPORT_FIELDS)
			
			for eachItem in report:
				ws.append([eachItem[field] for field in REPORT_FIELDS])
				
			wb.save(f"/tmp/{storeName}/{FN_REPORT}")
			
			reportName = f"REPORT - {storeName} - {TODAY_STRING}.xlsx"
			putToS3(reportName, f"/tmp/{storeName}/{FN_REPORT}")
		### End writing report ###
